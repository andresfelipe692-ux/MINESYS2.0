import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import re
import os
import openpyxl
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from tkcalendar import DateEntry
from PIL import Image, ImageTk, ImageDraw

# ══════════════════════════════════════════════════════════
#  FAVICON
# ══════════════════════════════════════════════════════════
def generar_favicon():
    ruta = os.path.join(os.path.dirname(os.path.abspath(__file__)), "favicon.ico")
    if not os.path.exists(ruta):
        img = Image.new("RGBA", (64, 64), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        draw.ellipse([2, 2, 62, 62], fill=(26, 58, 92, 255))
        draw.text((18, 18), "M", fill=(255, 255, 255, 255))
        img.save(ruta, format="ICO")
    return ruta

# ══════════════════════════════════════════════════════════
#  TEMAS
# ══════════════════════════════════════════════════════════
TEMAS = {
    "Claro": {
        "bg": "#f4f6f9",
        "fg": "#1a2b3c",
        "entry_bg": "#ffffff",
        "entry_fg": "#1a2b3c",
        "btn_bg": "#e0e0e0",
        "label_fg": "#2c3e50",
        "tabla_bg": "#ffffff",
        "tabla_fg": "#1a2b3c",
        "header_bg": "#1a3a5c",
        "header_fg": "#ffffff",
        "select_bg": "#1a3a5c",
        "select_fg": "#ffffff",
    },
    "Oscuro": {
        "bg": "#1a1e2e",
        "fg": "#e8ecf0",
        "entry_bg": "#2c3244",
        "entry_fg": "#e8ecf0",
        "btn_bg": "#2c3244",
        "label_fg": "#a8bcce",
        "tabla_bg": "#2a3244",
        "tabla_fg": "#e8ecf0",
        "header_bg": "#0d1520",
        "header_fg": "#7eb8f7",
        "select_bg": "#3a6ea5",
        "select_fg": "#ffffff",
    }
}

tema_actual = "Claro"
widgets_tema = []  # lista de (widget, tipo)

def registrar(widget, tipo):
    widgets_tema.append((widget, tipo))
    return widget

def aplicar_tema(nombre):
    global tema_actual
    tema_actual = nombre
    t = TEMAS[nombre]

    style = ttk.Style()
    style.theme_use("clam")
    style.configure("TNotebook", background=t["bg"])
    style.configure("TNotebook.Tab", background=t["btn_bg"], foreground=t["fg"],
                    padding=[10, 5], font=("Arial", 9, "bold"))
    style.map("TNotebook.Tab",
              background=[("selected", t["header_bg"])],
              foreground=[("selected", t["header_fg"])])
    style.configure("TFrame", background=t["bg"])
    style.configure("Treeview",
                    background=t["tabla_bg"],
                    foreground=t["tabla_fg"],
                    fieldbackground=t["tabla_bg"],
                    rowheight=22)
    style.configure("Treeview.Heading",
                    background=t["header_bg"],
                    foreground=t["header_fg"],
                    font=("Arial", 9, "bold"))
    style.map("Treeview",
              background=[("selected", t["select_bg"])],
              foreground=[("selected", t["select_fg"])])

    for widget, tipo in widgets_tema:
        try:
            if tipo == "bg":
                widget.configure(bg=t["bg"])
            elif tipo == "label":
                widget.configure(bg=t["bg"], fg=t["label_fg"])
            elif tipo == "title":
                widget.configure(bg=t["bg"], fg=t["header_bg"] if nombre == "Claro" else t["header_fg"])
            elif tipo == "entry":
                widget.configure(bg=t["entry_bg"], fg=t["entry_fg"],
                                 insertbackground=t["entry_fg"])
            elif tipo == "img_label":
                widget.configure(bg=t["entry_bg"])
        except Exception:
            pass

# ── Base de datos ──────────────────────────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "models", "minesys.db")

def conectar():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    return sqlite3.connect(DB_PATH)

def inicializar_db():
    con = conectar()
    cur = con.cursor()
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS yacimientos (
        codigo    TEXT PRIMARY KEY,
        nombre    TEXT NOT NULL,
        ubicacion TEXT NOT NULL,
        mineral   TEXT NOT NULL,
        metodo    TEXT NOT NULL,
        fecha     TEXT NOT NULL,
        reservas  INTEGER NOT NULL,
        vida      INTEGER NOT NULL,
        estado    TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS maquinaria (
        serie       TEXT PRIMARY KEY,
        tipo        TEXT NOT NULL,
        marca       TEXT NOT NULL,
        modelo      TEXT NOT NULL,
        capacidad   REAL NOT NULL,
        anio        INTEGER NOT NULL,
        horas       REAL NOT NULL,
        combustible TEXT NOT NULL,
        ubicacion   TEXT NOT NULL,
        estado      TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS empleados (
        id            TEXT PRIMARY KEY,
        nombre        TEXT NOT NULL,
        cargo         TEXT NOT NULL,
        edad          INTEGER NOT NULL,
        telefono      TEXT NOT NULL,
        correo        TEXT NOT NULL,
        direccion     TEXT NOT NULL,
        fecha_ingreso TEXT NOT NULL,
        salario       REAL NOT NULL
    );
    CREATE TABLE IF NOT EXISTS seguridad (
        id          TEXT PRIMARY KEY,
        zona        TEXT NOT NULL,
        riesgo      TEXT NOT NULL,
        descripcion TEXT NOT NULL,
        fecha       TEXT NOT NULL,
        responsable TEXT NOT NULL,
        accion      TEXT NOT NULL,
        estado      TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS auditoria (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        tabla     TEXT,
        operacion TEXT,
        fecha     TEXT DEFAULT (datetime('now')),
        detalle   TEXT
    );
    """)
    try:
        cur.execute("ALTER TABLE yacimientos ADD COLUMN imagen TEXT DEFAULT ''")
    except Exception:
        pass
    try:
        cur.execute("ALTER TABLE empleados ADD COLUMN imagen TEXT DEFAULT ''")
    except Exception:
        pass

    cur.execute("""CREATE TRIGGER IF NOT EXISTS sp_insert_yacimiento
    AFTER INSERT ON yacimientos BEGIN
        INSERT INTO auditoria(tabla, operacion, detalle)
        VALUES('yacimientos', 'INSERT', NEW.codigo || ' - ' || NEW.nombre);
    END""")
    cur.execute("""CREATE TRIGGER IF NOT EXISTS sp_update_yacimiento
    AFTER UPDATE ON yacimientos BEGIN
        INSERT INTO auditoria(tabla, operacion, detalle)
        VALUES('yacimientos', 'UPDATE', NEW.codigo || ' - ' || NEW.nombre);
    END""")
    cur.execute("""CREATE TRIGGER IF NOT EXISTS sp_delete_yacimiento
    AFTER DELETE ON yacimientos BEGIN
        INSERT INTO auditoria(tabla, operacion, detalle)
        VALUES('yacimientos', 'DELETE', OLD.codigo || ' - ' || OLD.nombre);
    END""")
    cur.execute("""CREATE TRIGGER IF NOT EXISTS sp_insert_maquinaria
    AFTER INSERT ON maquinaria BEGIN
        INSERT INTO auditoria(tabla, operacion, detalle)
        VALUES('maquinaria', 'INSERT', NEW.serie || ' - ' || NEW.tipo);
    END""")
    cur.execute("""CREATE TRIGGER IF NOT EXISTS sp_update_maquinaria
    AFTER UPDATE ON maquinaria BEGIN
        INSERT INTO auditoria(tabla, operacion, detalle)
        VALUES('maquinaria', 'UPDATE', NEW.serie || ' - ' || NEW.tipo);
    END""")
    cur.execute("""CREATE TRIGGER IF NOT EXISTS sp_delete_maquinaria
    AFTER DELETE ON maquinaria BEGIN
        INSERT INTO auditoria(tabla, operacion, detalle)
        VALUES('maquinaria', 'DELETE', OLD.serie || ' - ' || OLD.tipo);
    END""")
    cur.execute("""CREATE TRIGGER IF NOT EXISTS sp_insert_empleado
    AFTER INSERT ON empleados BEGIN
        INSERT INTO auditoria(tabla, operacion, detalle)
        VALUES('empleados', 'INSERT', NEW.id || ' - ' || NEW.nombre);
    END""")
    cur.execute("""CREATE TRIGGER IF NOT EXISTS sp_update_empleado
    AFTER UPDATE ON empleados BEGIN
        INSERT INTO auditoria(tabla, operacion, detalle)
        VALUES('empleados', 'UPDATE', NEW.id || ' - ' || NEW.nombre);
    END""")
    cur.execute("""CREATE TRIGGER IF NOT EXISTS sp_delete_empleado
    AFTER DELETE ON empleados BEGIN
        INSERT INTO auditoria(tabla, operacion, detalle)
        VALUES('empleados', 'DELETE', OLD.id || ' - ' || OLD.nombre);
    END""")
    cur.execute("""CREATE TRIGGER IF NOT EXISTS sp_insert_seguridad
    AFTER INSERT ON seguridad BEGIN
        INSERT INTO auditoria(tabla, operacion, detalle)
        VALUES('seguridad', 'INSERT', NEW.id || ' - ' || NEW.zona);
    END""")
    cur.execute("""CREATE TRIGGER IF NOT EXISTS sp_update_seguridad
    AFTER UPDATE ON seguridad BEGIN
        INSERT INTO auditoria(tabla, operacion, detalle)
        VALUES('seguridad', 'UPDATE', NEW.id || ' - ' || NEW.zona);
    END""")
    cur.execute("""CREATE TRIGGER IF NOT EXISTS sp_delete_seguridad
    AFTER DELETE ON seguridad BEGIN
        INSERT INTO auditoria(tabla, operacion, detalle)
        VALUES('seguridad', 'DELETE', OLD.id || ' - ' || OLD.zona);
    END""")
    con.commit()
    con.close()

# ── Ventana principal ──────────────────────────────────────────────────────
inicializar_db()

root = tk.Tk()
root.geometry("1050x680")
root.title("MineSys - Sistema Minero")

# Favicon
try:
    favicon_path = generar_favicon()
    root.iconbitmap(favicon_path)
except Exception:
    pass

# ── Barra superior con selector de tema ───────────────────────────────────
topbar = tk.Frame(root, height=45)
topbar.pack(side="top", fill="x")
registrar(topbar, "bg")

lbl_titulo = tk.Label(topbar, text="⛏  MineSys", font=("Arial", 14, "bold"), pady=8)
lbl_titulo.pack(side="left", padx=15)
registrar(lbl_titulo, "title")

frm_tema = tk.Frame(topbar)
frm_tema.pack(side="right", padx=15)
registrar(frm_tema, "bg")

tk.Label(frm_tema, text="Tema:", font=("Arial", 9)).pack(side="left", padx=4)

tema_var = tk.StringVar(value="Claro")

def cambiar_tema():
    aplicar_tema(tema_var.get())

for nombre in TEMAS:
    rb = tk.Radiobutton(frm_tema, text=nombre, variable=tema_var,
                        value=nombre, font=("Arial", 9), command=cambiar_tema)
    rb.pack(side="left", padx=3)
    registrar(rb, "label")

# ── Notebook ──────────────────────────────────────────────────────────────
notebook = ttk.Notebook(root)
tab1 = ttk.Frame(notebook)
tab2 = ttk.Frame(notebook)
tab3 = ttk.Frame(notebook)
tab4 = ttk.Frame(notebook)
notebook.add(tab1, text="⛰ Yacimientos")
notebook.add(tab2, text="🚜 Maquinaria")
notebook.add(tab3, text="👷 Empleados")
notebook.add(tab4, text="🦺 Seguridad")
notebook.pack(expand=True, fill="both", padx=6, pady=(0,6))

# ══════════════════════════════════════════════════════════
#  TAB 1 — YACIMIENTOS
# ══════════════════════════════════════════════════════════
campos_yacimiento = ["Código","Nombre","Ubicación","Mineral","Método","Fecha","Reservas","Vida","Estado"]
entradas_yacimiento = {}
img_path_yacimiento = {"ruta": ""}

lbl_yac = tk.Label(tab1, text="Registro de Yacimientos", font=("Arial",15,"bold"))
lbl_yac.pack(pady=8)
registrar(lbl_yac, "title")

form1 = tk.Frame(tab1)
form1.pack()
registrar(form1, "bg")

for i, campo in enumerate(campos_yacimiento):
    lbl = tk.Label(form1, text=campo, font=("Arial", 9))
    lbl.grid(row=i, column=0, padx=5, pady=2, sticky="w")
    registrar(lbl, "label")
    if campo == "Fecha":
        entrada = DateEntry(form1, date_pattern="yyyy-mm-dd")
    else:
        entrada = tk.Entry(form1)
        registrar(entrada, "entry")
    entrada.grid(row=i, column=1, padx=5, pady=2)
    entradas_yacimiento[campo] = entrada

img_preview_yac = tk.Label(form1, text="Sin imagen", width=15, height=7,
                            relief="groove", bg="#f0f0f0")
img_preview_yac.grid(row=0, column=2, rowspan=5, padx=10, pady=5)
registrar(img_preview_yac, "img_label")

def seleccionar_imagen_yacimiento():
    ruta = filedialog.askopenfilename(
        title="Seleccionar imagen",
        filetypes=[("Imágenes", "*.jpg *.jpeg *.png *.gif"), ("Todos", "*.*")]
    )
    if not ruta:
        return
    if os.path.splitext(ruta)[1].lower() not in [".jpg", ".jpeg", ".png", ".gif"]:
        messagebox.showerror("Error", "Solo se aceptan JPG, PNG o GIF.")
        return
    if os.path.getsize(ruta) / (1024 * 1024) > 5:
        messagebox.showerror("Error", "La imagen no puede superar 5 MB.")
        return
    img_path_yacimiento["ruta"] = ruta
    img = Image.open(ruta)
    img.thumbnail((120, 120))
    foto = ImageTk.PhotoImage(img)
    img_preview_yac.config(image=foto, text="")
    img_preview_yac.image = foto

def guardar_yacimiento():
    datos = []
    for c in campos_yacimiento:
        valor = entradas_yacimiento[c].get().strip()
        if c in ["Reservas", "Vida"]:
            if not valor.isdigit():
                messagebox.showerror("Error", f"'{c}' debe ser un número entero.")
                return
        if c in ["Nombre","Ubicación","Método","Estado"]:
            if len(valor) < 2:
                messagebox.showerror("Error", f"'{c}' es demasiado corto.")
                return
        datos.append(valor)
    if not messagebox.askyesno("Confirmar", "¿Desea guardar este yacimiento?"):
        return
    datos.append(img_path_yacimiento["ruta"])
    con = conectar()
    cur = con.cursor()
    try:
        cur.execute("INSERT INTO yacimientos VALUES (?,?,?,?,?,?,?,?,?,?)", datos)
        con.commit()
        messagebox.showinfo("OK", "Yacimiento guardado.")
        limpiar_yacimiento()
        cargar_yacimientos()
    except Exception as e:
        messagebox.showerror("Error", str(e))
    finally:
        con.close()

def cargar_yacimiento_seleccionado(event):
    sel = tabla_yacimientos.selection()
    if not sel:
        return
    vals = tabla_yacimientos.item(sel)["values"]
    limpiar_yacimiento()
    for i, c in enumerate(campos_yacimiento):
        entradas_yacimiento[c].insert(0, str(vals[i]))
    if len(vals) > len(campos_yacimiento):
        ruta = str(vals[len(campos_yacimiento)])
        img_path_yacimiento["ruta"] = ruta
        if ruta and os.path.exists(ruta):
            img = Image.open(ruta)
            img.thumbnail((120, 120))
            foto = ImageTk.PhotoImage(img)
            img_preview_yac.config(image=foto, text="")
            img_preview_yac.image = foto

def actualizar_yacimiento():
    sel = tabla_yacimientos.selection()
    if not sel:
        messagebox.showwarning("Aviso", "Seleccione un registro para actualizar.")
        return
    if not messagebox.askyesno("Confirmar", "¿Desea actualizar este yacimiento?"):
        return
    codigo_original = tabla_yacimientos.item(sel)["values"][0]
    datos = []
    for c in campos_yacimiento:
        valor = entradas_yacimiento[c].get().strip()
        if c in ["Reservas", "Vida"]:
            if not valor.isdigit():
                messagebox.showerror("Error", f"'{c}' debe ser un número.")
                return
        datos.append(valor)
    datos.append(img_path_yacimiento["ruta"])
    datos.append(codigo_original)
    con = conectar()
    cur = con.cursor()
    try:
        cur.execute("""UPDATE yacimientos SET codigo=?,nombre=?,ubicacion=?,
            mineral=?,metodo=?,fecha=?,reservas=?,vida=?,estado=?,imagen=?
            WHERE codigo=?""", datos)
        con.commit()
        messagebox.showinfo("OK", "Yacimiento actualizado.")
        limpiar_yacimiento()
        cargar_yacimientos()
    except Exception as e:
        messagebox.showerror("Error", str(e))
    finally:
        con.close()

def limpiar_yacimiento():
    for e in entradas_yacimiento.values():
        try:
            e.delete(0, "end")
        except Exception:
            pass
    img_preview_yac.config(image="", text="Sin imagen")
    img_path_yacimiento["ruta"] = ""

def eliminar_yacimiento():
    sel = tabla_yacimientos.selection()
    if not sel:
        messagebox.showwarning("Aviso", "Seleccione un registro.")
        return
    if not messagebox.askyesno("Confirmar", "¿Está seguro de eliminar este yacimiento?"):
        return
    codigo = tabla_yacimientos.item(sel)["values"][0]
    con = conectar()
    cur = con.cursor()
    cur.execute("DELETE FROM yacimientos WHERE codigo=?", (codigo,))
    con.commit()
    con.close()
    cargar_yacimientos()

def exportar_yacimientos_excel():
    con = conectar()
    cur = con.cursor()
    cur.execute("SELECT codigo,nombre,ubicacion,mineral,metodo,fecha,reservas,vida,estado FROM yacimientos")
    filas = cur.fetchall()
    con.close()
    archivo = filedialog.asksaveasfilename(defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")], title="Guardar Excel")
    if not archivo:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yacimientos"
    ws.append(campos_yacimiento)
    for fila in filas:
        ws.append(list(fila))
    wb.save(archivo)
    messagebox.showinfo("OK", "Excel guardado.")

def exportar_yacimientos_pdf():
    con = conectar()
    cur = con.cursor()
    cur.execute("SELECT codigo,nombre,ubicacion,mineral,metodo,fecha,reservas,vida,estado FROM yacimientos")
    filas = cur.fetchall()
    con.close()
    archivo = filedialog.asksaveasfilename(defaultextension=".pdf",
        filetypes=[("PDF", "*.pdf")], title="Guardar PDF")
    if not archivo:
        return
    doc = SimpleDocTemplate(archivo, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elementos = [Paragraph("Reporte de Yacimientos", styles["Title"]), Spacer(1, 0.5*cm)]
    data = [campos_yacimiento] + [list(f) for f in filas]
    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  colors.HexColor("#1a3a5c")),
        ("TEXTCOLOR",     (0,0), (-1,0),  colors.white),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.HexColor("#ebf3fb"), colors.white]),
        ("GRID",          (0,0), (-1,-1), 0.5, colors.grey),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    elementos.append(tabla)
    doc.build(elementos)
    messagebox.showinfo("OK", "PDF guardado.")

btnframe1 = tk.Frame(tab1)
btnframe1.pack(pady=6)
registrar(btnframe1, "bg")
tk.Button(btnframe1, text="💾 Guardar",       bg="green",   fg="white", command=guardar_yacimiento).pack(side="left", padx=3)
tk.Button(btnframe1, text="✏ Actualizar",    bg="blue",    fg="white", command=actualizar_yacimiento).pack(side="left", padx=3)
tk.Button(btnframe1, text="🔄 Limpiar",       bg="orange",              command=limpiar_yacimiento).pack(side="left", padx=3)
tk.Button(btnframe1, text="🗑 Eliminar",      bg="red",     fg="white", command=eliminar_yacimiento).pack(side="left", padx=3)
tk.Button(btnframe1, text="📊 Excel",         bg="#1a5c8a", fg="white", command=exportar_yacimientos_excel).pack(side="left", padx=3)
tk.Button(btnframe1, text="📄 PDF",           bg="#8a1a1a", fg="white", command=exportar_yacimientos_pdf).pack(side="left", padx=3)
tk.Button(btnframe1, text="🖼 Imagen",        bg="#6a3a8a", fg="white", command=seleccionar_imagen_yacimiento).pack(side="left", padx=3)

tabla_yacimientos = ttk.Treeview(tab1, columns=campos_yacimiento, show="headings", height=6)
for c in campos_yacimiento:
    tabla_yacimientos.heading(c, text=c)
    tabla_yacimientos.column(c, width=100)
tabla_yacimientos.pack(expand=True, fill="both", padx=8, pady=4)
tabla_yacimientos.bind("<<TreeviewSelect>>", cargar_yacimiento_seleccionado)

def cargar_yacimientos():
    con = conectar()
    cur = con.cursor()
    tabla_yacimientos.delete(*tabla_yacimientos.get_children())
    cur.execute("SELECT codigo,nombre,ubicacion,mineral,metodo,fecha,reservas,vida,estado,imagen FROM yacimientos")
    for fila in cur.fetchall():
        tabla_yacimientos.insert("", tk.END, values=fila)
    con.close()

# ══════════════════════════════════════════════════════════
#  TAB 2 — MAQUINARIA
# ══════════════════════════════════════════════════════════
campos_maquinaria = ["Serie","Tipo","Marca","Modelo","Capacidad","Año","Horas","Combustible","Ubicación","Estado"]
entradas_maquinaria = {}

lbl_maq = tk.Label(tab2, text="Registro de Maquinaria", font=("Arial",15,"bold"))
lbl_maq.pack(pady=8)
registrar(lbl_maq, "title")

form2 = tk.Frame(tab2)
form2.pack()
registrar(form2, "bg")

for i, campo in enumerate(campos_maquinaria):
    lbl = tk.Label(form2, text=campo, font=("Arial", 9))
    lbl.grid(row=i, column=0, padx=5, pady=2, sticky="w")
    registrar(lbl, "label")
    entrada = tk.Entry(form2)
    entrada.grid(row=i, column=1, padx=5, pady=2)
    registrar(entrada, "entry")
    entradas_maquinaria[campo] = entrada

def guardar_maquinaria():
    datos = []
    for c in campos_maquinaria:
        valor = entradas_maquinaria[c].get().strip()
        if c in ["Capacidad","Horas","Año"]:
            try:
                float(valor)
            except ValueError:
                messagebox.showerror("Error", f"'{c}' debe ser un número.")
                return
        if c in ["Tipo","Marca","Modelo","Combustible","Ubicación","Estado"]:
            if len(valor) < 2:
                messagebox.showerror("Error", f"'{c}' es demasiado corto.")
                return
        datos.append(valor)
    if not messagebox.askyesno("Confirmar", "¿Desea guardar esta maquinaria?"):
        return
    con = conectar()
    cur = con.cursor()
    try:
        cur.execute("INSERT INTO maquinaria VALUES (?,?,?,?,?,?,?,?,?,?)", datos)
        con.commit()
        messagebox.showinfo("OK", "Maquinaria guardada.")
        limpiar_maquinaria()
        cargar_maquinaria()
    except Exception as e:
        messagebox.showerror("Error", str(e))
    finally:
        con.close()

def cargar_maquinaria_seleccionada(event):
    sel = tabla_maquinaria.selection()
    if not sel:
        return
    vals = tabla_maquinaria.item(sel)["values"]
    limpiar_maquinaria()
    for i, c in enumerate(campos_maquinaria):
        entradas_maquinaria[c].insert(0, str(vals[i]))

def actualizar_maquinaria():
    sel = tabla_maquinaria.selection()
    if not sel:
        messagebox.showwarning("Aviso", "Seleccione un registro para actualizar.")
        return
    if not messagebox.askyesno("Confirmar", "¿Desea actualizar esta maquinaria?"):
        return
    serie_original = tabla_maquinaria.item(sel)["values"][0]
    datos = [entradas_maquinaria[c].get().strip() for c in campos_maquinaria]
    datos.append(serie_original)
    con = conectar()
    cur = con.cursor()
    try:
        cur.execute("""UPDATE maquinaria SET serie=?,tipo=?,marca=?,modelo=?,
            capacidad=?,anio=?,horas=?,combustible=?,ubicacion=?,estado=?
            WHERE serie=?""", datos)
        con.commit()
        messagebox.showinfo("OK", "Maquinaria actualizada.")
        limpiar_maquinaria()
        cargar_maquinaria()
    except Exception as e:
        messagebox.showerror("Error", str(e))
    finally:
        con.close()

def limpiar_maquinaria():
    for e in entradas_maquinaria.values():
        e.delete(0, "end")

def eliminar_maquinaria():
    sel = tabla_maquinaria.selection()
    if not sel:
        messagebox.showwarning("Aviso", "Seleccione un registro.")
        return
    if not messagebox.askyesno("Confirmar", "¿Está seguro de eliminar esta maquinaria?"):
        return
    serie = tabla_maquinaria.item(sel)["values"][0]
    con = conectar()
    cur = con.cursor()
    cur.execute("DELETE FROM maquinaria WHERE serie=?", (serie,))
    con.commit()
    con.close()
    cargar_maquinaria()

def exportar_maquinaria_excel():
    con = conectar()
    cur = con.cursor()
    cur.execute("SELECT * FROM maquinaria")
    filas = cur.fetchall()
    con.close()
    archivo = filedialog.asksaveasfilename(defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")], title="Guardar Excel")
    if not archivo:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maquinaria"
    ws.append(campos_maquinaria)
    for fila in filas:
        ws.append(list(fila))
    wb.save(archivo)
    messagebox.showinfo("OK", "Excel guardado.")

def exportar_maquinaria_pdf():
    con = conectar()
    cur = con.cursor()
    cur.execute("SELECT * FROM maquinaria")
    filas = cur.fetchall()
    con.close()
    archivo = filedialog.asksaveasfilename(defaultextension=".pdf",
        filetypes=[("PDF", "*.pdf")], title="Guardar PDF")
    if not archivo:
        return
    doc = SimpleDocTemplate(archivo, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elementos = [Paragraph("Reporte de Maquinaria", styles["Title"]), Spacer(1, 0.5*cm)]
    data = [campos_maquinaria] + [list(f) for f in filas]
    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  colors.HexColor("#1a3a5c")),
        ("TEXTCOLOR",     (0,0), (-1,0),  colors.white),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.HexColor("#ebf3fb"), colors.white]),
        ("GRID",          (0,0), (-1,-1), 0.5, colors.grey),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    elementos.append(tabla)
    doc.build(elementos)
    messagebox.showinfo("OK", "PDF guardado.")

btnframe2 = tk.Frame(tab2)
btnframe2.pack(pady=6)
registrar(btnframe2, "bg")
tk.Button(btnframe2, text="💾 Guardar",    bg="green",   fg="white", command=guardar_maquinaria).pack(side="left", padx=3)
tk.Button(btnframe2, text="✏ Actualizar", bg="blue",    fg="white", command=actualizar_maquinaria).pack(side="left", padx=3)
tk.Button(btnframe2, text="🔄 Limpiar",    bg="orange",              command=limpiar_maquinaria).pack(side="left", padx=3)
tk.Button(btnframe2, text="🗑 Eliminar",   bg="red",     fg="white", command=eliminar_maquinaria).pack(side="left", padx=3)
tk.Button(btnframe2, text="📊 Excel",      bg="#1a5c8a", fg="white", command=exportar_maquinaria_excel).pack(side="left", padx=3)
tk.Button(btnframe2, text="📄 PDF",        bg="#8a1a1a", fg="white", command=exportar_maquinaria_pdf).pack(side="left", padx=3)

tabla_maquinaria = ttk.Treeview(tab2, columns=campos_maquinaria, show="headings", height=6)
for c in campos_maquinaria:
    tabla_maquinaria.heading(c, text=c)
    tabla_maquinaria.column(c, width=100)
tabla_maquinaria.pack(expand=True, fill="both", padx=8, pady=4)
tabla_maquinaria.bind("<<TreeviewSelect>>", cargar_maquinaria_seleccionada)

def cargar_maquinaria():
    con = conectar()
    cur = con.cursor()
    tabla_maquinaria.delete(*tabla_maquinaria.get_children())
    cur.execute("SELECT * FROM maquinaria")
    for fila in cur.fetchall():
        tabla_maquinaria.insert("", tk.END, values=fila)
    con.close()

# ══════════════════════════════════════════════════════════
#  TAB 3 — EMPLEADOS
# ══════════════════════════════════════════════════════════
campos_empleados = ["ID","Nombre","Cargo","Edad","Teléfono","Correo","Dirección","Fecha ingreso","Salario"]
entradas_empleados = {}
img_path_empleado = {"ruta": ""}

lbl_emp = tk.Label(tab3, text="Registro de Empleados", font=("Arial",15,"bold"))
lbl_emp.pack(pady=8)
registrar(lbl_emp, "title")

form3 = tk.Frame(tab3)
form3.pack()
registrar(form3, "bg")

for i, campo in enumerate(campos_empleados):
    lbl = tk.Label(form3, text=campo, font=("Arial", 9))
    lbl.grid(row=i, column=0, padx=5, pady=2, sticky="w")
    registrar(lbl, "label")
    if campo == "Fecha ingreso":
        entrada = DateEntry(form3, date_pattern="yyyy-mm-dd")
    else:
        entrada = tk.Entry(form3)
        registrar(entrada, "entry")
    entrada.grid(row=i, column=1, padx=5, pady=2)
    entradas_empleados[campo] = entrada

img_preview_emp = tk.Label(form3, text="Sin imagen", width=15, height=7,
                            relief="groove", bg="#f0f0f0")
img_preview_emp.grid(row=0, column=2, rowspan=5, padx=10, pady=5)
registrar(img_preview_emp, "img_label")

def seleccionar_imagen_empleado():
    ruta = filedialog.askopenfilename(
        title="Seleccionar imagen",
        filetypes=[("Imágenes", "*.jpg *.jpeg *.png *.gif"), ("Todos", "*.*")]
    )
    if not ruta:
        return
    if os.path.splitext(ruta)[1].lower() not in [".jpg", ".jpeg", ".png", ".gif"]:
        messagebox.showerror("Error", "Solo se aceptan JPG, PNG o GIF.")
        return
    if os.path.getsize(ruta) / (1024 * 1024) > 5:
        messagebox.showerror("Error", "La imagen no puede superar 5 MB.")
        return
    img_path_empleado["ruta"] = ruta
    img = Image.open(ruta)
    img.thumbnail((120, 120))
    foto = ImageTk.PhotoImage(img)
    img_preview_emp.config(image=foto, text="")
    img_preview_emp.image = foto

def guardar_empleado():
    datos = []
    for c in campos_empleados:
        valor = entradas_empleados[c].get().strip()
        if c in ["Edad", "Salario"]:
            try:
                float(valor)
            except ValueError:
                messagebox.showerror("Error", f"'{c}' debe ser un número.")
                return
        if c == "Correo":
            if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", valor):
                messagebox.showerror("Error", "El formato del correo es inválido.")
                return
        if c in ["Nombre","Cargo","Teléfono","Dirección"]:
            if len(valor) < 2:
                messagebox.showerror("Error", f"'{c}' es demasiado corto.")
                return
        datos.append(valor)
    if not messagebox.askyesno("Confirmar", "¿Desea guardar este empleado?"):
        return
    datos.append(img_path_empleado["ruta"])
    con = conectar()
    cur = con.cursor()
    try:
        cur.execute("INSERT INTO empleados VALUES (?,?,?,?,?,?,?,?,?,?)", datos)
        con.commit()
        messagebox.showinfo("OK", "Empleado guardado.")
        limpiar_empleado()
        cargar_empleados()
    except Exception as e:
        messagebox.showerror("Error", str(e))
    finally:
        con.close()

def cargar_empleado_seleccionado(event):
    sel = tabla_empleados.selection()
    if not sel:
        return
    vals = tabla_empleados.item(sel)["values"]
    limpiar_empleado()
    for i, c in enumerate(campos_empleados):
        try:
            entradas_empleados[c].insert(0, str(vals[i]))
        except Exception:
            pass
    if len(vals) > len(campos_empleados):
        ruta = str(vals[len(campos_empleados)])
        img_path_empleado["ruta"] = ruta
        if ruta and os.path.exists(ruta):
            img = Image.open(ruta)
            img.thumbnail((120, 120))
            foto = ImageTk.PhotoImage(img)
            img_preview_emp.config(image=foto, text="")
            img_preview_emp.image = foto

def actualizar_empleado():
    sel = tabla_empleados.selection()
    if not sel:
        messagebox.showwarning("Aviso", "Seleccione un registro para actualizar.")
        return
    if not messagebox.askyesno("Confirmar", "¿Desea actualizar este empleado?"):
        return
    id_original = tabla_empleados.item(sel)["values"][0]
    datos = []
    for c in campos_empleados:
        valor = entradas_empleados[c].get().strip()
        if c == "Correo":
            if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", valor):
                messagebox.showerror("Error", "El formato del correo es inválido.")
                return
        datos.append(valor)
    datos.append(img_path_empleado["ruta"])
    datos.append(id_original)
    con = conectar()
    cur = con.cursor()
    try:
        cur.execute("""UPDATE empleados SET id=?,nombre=?,cargo=?,edad=?,
            telefono=?,correo=?,direccion=?,fecha_ingreso=?,salario=?,imagen=?
            WHERE id=?""", datos)
        con.commit()
        messagebox.showinfo("OK", "Empleado actualizado.")
        limpiar_empleado()
        cargar_empleados()
    except Exception as e:
        messagebox.showerror("Error", str(e))
    finally:
        con.close()

def limpiar_empleado():
    for e in entradas_empleados.values():
        try:
            e.delete(0, "end")
        except Exception:
            pass
    img_preview_emp.config(image="", text="Sin imagen")
    img_path_empleado["ruta"] = ""

def eliminar_empleado():
    sel = tabla_empleados.selection()
    if not sel:
        messagebox.showwarning("Aviso", "Seleccione un registro.")
        return
    if not messagebox.askyesno("Confirmar", "¿Está seguro de eliminar este empleado?"):
        return
    idemp = tabla_empleados.item(sel)["values"][0]
    con = conectar()
    cur = con.cursor()
    cur.execute("DELETE FROM empleados WHERE id=?", (idemp,))
    con.commit()
    con.close()
    cargar_empleados()

def exportar_empleados_excel():
    con = conectar()
    cur = con.cursor()
    cur.execute("SELECT id,nombre,cargo,edad,telefono,correo,direccion,fecha_ingreso,salario FROM empleados")
    filas = cur.fetchall()
    con.close()
    archivo = filedialog.asksaveasfilename(defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")], title="Guardar Excel")
    if not archivo:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empleados"
    ws.append(campos_empleados)
    for fila in filas:
        ws.append(list(fila))
    wb.save(archivo)
    messagebox.showinfo("OK", "Excel guardado.")

def exportar_empleados_pdf():
    con = conectar()
    cur = con.cursor()
    cur.execute("SELECT id,nombre,cargo,edad,telefono,correo,direccion,fecha_ingreso,salario FROM empleados")
    filas = cur.fetchall()
    con.close()
    archivo = filedialog.asksaveasfilename(defaultextension=".pdf",
        filetypes=[("PDF", "*.pdf")], title="Guardar PDF")
    if not archivo:
        return
    doc = SimpleDocTemplate(archivo, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elementos = [Paragraph("Reporte de Empleados", styles["Title"]), Spacer(1, 0.5*cm)]
    data = [campos_empleados] + [list(f) for f in filas]
    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  colors.HexColor("#1a3a5c")),
        ("TEXTCOLOR",     (0,0), (-1,0),  colors.white),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.HexColor("#ebf3fb"), colors.white]),
        ("GRID",          (0,0), (-1,-1), 0.5, colors.grey),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    elementos.append(tabla)
    doc.build(elementos)
    messagebox.showinfo("OK", "PDF guardado.")

btnframe3 = tk.Frame(tab3)
btnframe3.pack(pady=6)
registrar(btnframe3, "bg")
tk.Button(btnframe3, text="💾 Guardar",    bg="green",   fg="white", command=guardar_empleado).pack(side="left", padx=3)
tk.Button(btnframe3, text="✏ Actualizar", bg="blue",    fg="white", command=actualizar_empleado).pack(side="left", padx=3)
tk.Button(btnframe3, text="🔄 Limpiar",    bg="orange",              command=limpiar_empleado).pack(side="left", padx=3)
tk.Button(btnframe3, text="🗑 Eliminar",   bg="red",     fg="white", command=eliminar_empleado).pack(side="left", padx=3)
tk.Button(btnframe3, text="📊 Excel",      bg="#1a5c8a", fg="white", command=exportar_empleados_excel).pack(side="left", padx=3)
tk.Button(btnframe3, text="📄 PDF",        bg="#8a1a1a", fg="white", command=exportar_empleados_pdf).pack(side="left", padx=3)
tk.Button(btnframe3, text="🖼 Foto",       bg="#6a3a8a", fg="white", command=seleccionar_imagen_empleado).pack(side="left", padx=3)

tabla_empleados = ttk.Treeview(tab3, columns=campos_empleados, show="headings", height=6)
for c in campos_empleados:
    tabla_empleados.heading(c, text=c)
    tabla_empleados.column(c, width=100)
tabla_empleados.pack(expand=True, fill="both", padx=8, pady=4)
tabla_empleados.bind("<<TreeviewSelect>>", cargar_empleado_seleccionado)

def cargar_empleados():
    con = conectar()
    cur = con.cursor()
    tabla_empleados.delete(*tabla_empleados.get_children())
    cur.execute("SELECT id,nombre,cargo,edad,telefono,correo,direccion,fecha_ingreso,salario,imagen FROM empleados")
    for fila in cur.fetchall():
        tabla_empleados.insert("", tk.END, values=fila)
    con.close()

# ══════════════════════════════════════════════════════════
#  TAB 4 — SEGURIDAD
# ══════════════════════════════════════════════════════════
campos_seguridad = ["ID","Zona","Riesgo","Descripción","Fecha","Responsable","Acción","Estado"]
entradas_seguridad = {}

lbl_seg = tk.Label(tab4, text="Registro de Seguridad", font=("Arial",15,"bold"))
lbl_seg.pack(pady=8)
registrar(lbl_seg, "title")

form4 = tk.Frame(tab4)
form4.pack()
registrar(form4, "bg")

for i, campo in enumerate(campos_seguridad):
    lbl = tk.Label(form4, text=campo, font=("Arial", 9))
    lbl.grid(row=i, column=0, padx=5, pady=2, sticky="w")
    registrar(lbl, "label")
    if campo == "Fecha":
        entrada = DateEntry(form4, date_pattern="yyyy-mm-dd")
    else:
        entrada = tk.Entry(form4)
        registrar(entrada, "entry")
    entrada.grid(row=i, column=1, padx=5, pady=2)
    entradas_seguridad[campo] = entrada

def guardar_seguridad():
    datos = []
    for c in campos_seguridad:
        valor = entradas_seguridad[c].get().strip()
        if c in ["Zona","Riesgo","Acción","Estado","Responsable"]:
            if len(valor) < 2:
                messagebox.showerror("Error", f"'{c}' es demasiado corto.")
                return
        datos.append(valor)
    if not messagebox.askyesno("Confirmar", "¿Desea guardar este registro?"):
        return
    con = conectar()
    cur = con.cursor()
    try:
        cur.execute("INSERT INTO seguridad VALUES (?,?,?,?,?,?,?,?)", datos)
        con.commit()
        messagebox.showinfo("OK", "Registro guardado.")
        limpiar_seguridad()
        cargar_seguridad()
    except Exception as e:
        messagebox.showerror("Error", str(e))
    finally:
        con.close()

def cargar_seguridad_seleccionada(event):
    sel = tabla_seguridad.selection()
    if not sel:
        return
    vals = tabla_seguridad.item(sel)["values"]
    limpiar_seguridad()
    for i, c in enumerate(campos_seguridad):
        try:
            entradas_seguridad[c].insert(0, str(vals[i]))
        except Exception:
            pass

def actualizar_seguridad():
    sel = tabla_seguridad.selection()
    if not sel:
        messagebox.showwarning("Aviso", "Seleccione un registro para actualizar.")
        return
    if not messagebox.askyesno("Confirmar", "¿Desea actualizar este registro?"):
        return
    id_original = tabla_seguridad.item(sel)["values"][0]
    datos = [entradas_seguridad[c].get().strip() for c in campos_seguridad]
    datos.append(id_original)
    con = conectar()
    cur = con.cursor()
    try:
        cur.execute("""UPDATE seguridad SET id=?,zona=?,riesgo=?,descripcion=?,
            fecha=?,responsable=?,accion=?,estado=? WHERE id=?""", datos)
        con.commit()
        messagebox.showinfo("OK", "Registro actualizado.")
        limpiar_seguridad()
        cargar_seguridad()
    except Exception as e:
        messagebox.showerror("Error", str(e))
    finally:
        con.close()

def limpiar_seguridad():
    for e in entradas_seguridad.values():
        try:
            e.delete(0, "end")
        except Exception:
            pass

def eliminar_seguridad():
    sel = tabla_seguridad.selection()
    if not sel:
        messagebox.showwarning("Aviso", "Seleccione un registro.")
        return
    if not messagebox.askyesno("Confirmar", "¿Está seguro de eliminar este registro?"):
        return
    idseg = tabla_seguridad.item(sel)["values"][0]
    con = conectar()
    cur = con.cursor()
    cur.execute("DELETE FROM seguridad WHERE id=?", (idseg,))
    con.commit()
    con.close()
    cargar_seguridad()

def exportar_seguridad_excel():
    con = conectar()
    cur = con.cursor()
    cur.execute("SELECT * FROM seguridad")
    filas = cur.fetchall()
    con.close()
    archivo = filedialog.asksaveasfilename(defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")], title="Guardar Excel")
    if not archivo:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Seguridad"
    ws.append(campos_seguridad)
    for fila in filas:
        ws.append(list(fila))
    wb.save(archivo)
    messagebox.showinfo("OK", "Excel guardado.")

def exportar_seguridad_pdf():
    con = conectar()
    cur = con.cursor()
    cur.execute("SELECT * FROM seguridad")
    filas = cur.fetchall()
    con.close()
    archivo = filedialog.asksaveasfilename(defaultextension=".pdf",
        filetypes=[("PDF", "*.pdf")], title="Guardar PDF")
    if not archivo:
        return
    doc = SimpleDocTemplate(archivo, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elementos = [Paragraph("Reporte de Seguridad", styles["Title"]), Spacer(1, 0.5*cm)]
    data = [campos_seguridad] + [list(f) for f in filas]
    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  colors.HexColor("#1a3a5c")),
        ("TEXTCOLOR",     (0,0), (-1,0),  colors.white),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.HexColor("#ebf3fb"), colors.white]),
        ("GRID",          (0,0), (-1,-1), 0.5, colors.grey),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    elementos.append(tabla)
    doc.build(elementos)
    messagebox.showinfo("OK", "PDF guardado.")

btnframe4 = tk.Frame(tab4)
btnframe4.pack(pady=6)
registrar(btnframe4, "bg")
tk.Button(btnframe4, text="💾 Guardar",    bg="green",   fg="white", command=guardar_seguridad).pack(side="left", padx=3)
tk.Button(btnframe4, text="✏ Actualizar", bg="blue",    fg="white", command=actualizar_seguridad).pack(side="left", padx=3)
tk.Button(btnframe4, text="🔄 Limpiar",    bg="orange",              command=limpiar_seguridad).pack(side="left", padx=3)
tk.Button(btnframe4, text="🗑 Eliminar",   bg="red",     fg="white", command=eliminar_seguridad).pack(side="left", padx=3)
tk.Button(btnframe4, text="📊 Excel",      bg="#1a5c8a", fg="white", command=exportar_seguridad_excel).pack(side="left", padx=3)
tk.Button(btnframe4, text="📄 PDF",        bg="#8a1a1a", fg="white", command=exportar_seguridad_pdf).pack(side="left", padx=3)

tabla_seguridad = ttk.Treeview(tab4, columns=campos_seguridad, show="headings", height=6)
for c in campos_seguridad:
    tabla_seguridad.heading(c, text=c)
    tabla_seguridad.column(c, width=120)
tabla_seguridad.pack(expand=True, fill="both", padx=8, pady=4)
tabla_seguridad.bind("<<TreeviewSelect>>", cargar_seguridad_seleccionada)

def cargar_seguridad():
    con = conectar()
    cur = con.cursor()
    tabla_seguridad.delete(*tabla_seguridad.get_children())
    cur.execute("SELECT * FROM seguridad")
    for fila in cur.fetchall():
        tabla_seguridad.insert("", tk.END, values=fila)
    con.close()

# ── Cargar datos y aplicar tema inicial ───────────────────────────────────
cargar_yacimientos()
cargar_maquinaria()
cargar_empleados()
cargar_seguridad()
aplicar_tema("Claro")

root.mainloop()